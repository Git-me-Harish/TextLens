import io
import csv
import json
from typing import Any


def _flatten(obj: Any, prefix: str = "", sep: str = ".") -> dict:
    """Recursively flatten nested JSON into dot-notation keys."""
    items = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}{sep}{k}" if prefix else k
            if isinstance(v, (dict, list)):
                items.update(_flatten(v, key, sep))
            else:
                items[key] = v
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            key = f"{prefix}[{i}]"
            if isinstance(v, (dict, list)):
                items.update(_flatten(v, key, sep))
            else:
                items[key] = v
    else:
        items[prefix] = obj
    return items


def to_csv(structured_result: dict) -> bytes:
    """Convert structured agent result to CSV bytes."""
    output = io.StringIO()

    # If result has a list (e.g. transactions, line_items) — use that as rows
    list_key = next((k for k, v in structured_result.items() if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict)), None)

    if list_key:
        rows = structured_result[list_key]
        writer = csv.DictWriter(output, fieldnames=rows[0].keys(), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    else:
        # Fallback: flatten entire result as two-column key/value
        flat = _flatten(structured_result)
        writer = csv.writer(output)
        writer.writerow(["Field", "Value"])
        for k, v in flat.items():
            writer.writerow([k, v])

    return output.getvalue().encode("utf-8")


def to_excel(structured_result: dict, pipeline_type: str) -> bytes:
    """Convert structured agent result to Excel bytes."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="E2E1DA"),
        right=Side(style="thin", color="E2E1DA"),
    )
    alt_fill = PatternFill("solid", fgColor="F3F2EF")

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    def style_row(cell, alt=False):
        if alt:
            cell.fill = alt_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border

    # Main summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.row_dimensions[1].height = 28

    flat = _flatten(structured_result)
    # Skip list fields for summary
    flat_simple = {k: v for k, v in flat.items() if not isinstance(v, (dict, list))}

    ws.append(["Field", "Value"])
    style_header(ws["A1"])
    style_header(ws["B1"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52

    for i, (k, v) in enumerate(flat_simple.items()):
        ws.append([k.replace("_", " ").title(), str(v) if v is not None else "—"])
        style_row(ws.cell(i + 2, 1), alt=(i % 2 == 1))
        style_row(ws.cell(i + 2, 2), alt=(i % 2 == 1))

    # Detail sheets for list fields
    for key, val in structured_result.items():
        if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
            ws2 = wb.create_sheet(title=key.replace("_", " ").title()[:31])
            headers = list(val[0].keys())
            ws2.append(headers)
            for j, h in enumerate(headers, 1):
                style_header(ws2.cell(1, j))
                ws2.column_dimensions[get_column_letter(j)].width = 22

            for i, row in enumerate(val):
                ws2.append([str(row.get(h, "")) for h in headers])
                for j in range(1, len(headers) + 1):
                    style_row(ws2.cell(i + 2, j), alt=(i % 2 == 1))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
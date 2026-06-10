"""
Batch processing routes.

POST /api/batch/upload   — upload multiple files + pick pipeline → creates BatchJob
GET  /api/batch          — list user's batch jobs
GET  /api/batch/{id}     — poll batch job status + item progress
GET  /api/batch/{id}/results  — download all structured results as Excel (one sheet per file)
DELETE /api/batch/{id}   — cancel/delete
"""
import asyncio
import logging
import os
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, BackgroundTasks
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.api.deps import get_current_user
from app.core.config import settings
from app.db.database import get_db
from app.models.models import User, BatchJob, BatchItem, BatchStatus, AgentDomain
from app.schemas.schemas import BatchJobCreate, BatchJobOut, BatchJobListResponse
from app.services.batch_service import run_batch_job
from app.services.agent_service import get_pipeline_catalog

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/batch", tags=["batch"])

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".tiff", ".tif", ".bmp"}
MAX_FILES_PER_BATCH = 50


def _validate_pipeline(domain: str, pipeline_type: str) -> None:
    """Raise 400 if domain/pipeline combo is unknown."""
    catalog = get_pipeline_catalog()
    if domain not in catalog:
        raise HTTPException(status_code=400, detail=f"Unknown domain: {domain}")
    if pipeline_type not in catalog[domain]["pipelines"]:
        raise HTTPException(status_code=400, detail=f"Unknown pipeline: {pipeline_type} for domain: {domain}")


def _extract_files_from_upload(
    files: list[UploadFile],
    upload_dir: Path,
) -> list[tuple[str, str]]:
    """
    Save uploaded files to disk, handling ZIP expansion.
    Returns list of (file_path, original_filename).
    """
    saved: list[tuple[str, str]] = []

    for upload in files:
        content = upload.file.read()
        ext = Path(upload.filename or "file").suffix.lower()

        if ext == ".zip":
            # Expand ZIP — process each file inside
            tmp_zip = upload_dir / f"{uuid.uuid4()}.zip"
            tmp_zip.write_bytes(content)
            try:
                with zipfile.ZipFile(tmp_zip, "r") as zf:
                    for member in zf.namelist():
                        member_ext = Path(member).suffix.lower()
                        if member_ext not in ALLOWED_EXTENSIONS:
                            continue
                        member_bytes = zf.read(member)
                        out_path = upload_dir / f"{uuid.uuid4()}{member_ext}"
                        out_path.write_bytes(member_bytes)
                        saved.append((str(out_path), Path(member).name))
            finally:
                tmp_zip.unlink(missing_ok=True)
        elif ext in ALLOWED_EXTENSIONS:
            out_path = upload_dir / f"{uuid.uuid4()}{ext}"
            out_path.write_bytes(content)
            saved.append((str(out_path), upload.filename or "file"))

    return saved


@router.post("", response_model=BatchJobOut, status_code=202)
async def create_batch_job(
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    name: str = Form(default="Batch Job"),
    domain: str = Form(...),
    pipeline_type: str = Form(...),
    user_instructions: str = Form(default=""),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Upload up to 50 files (or a ZIP) + select domain pipeline.
    Returns BatchJob immediately — processing runs in background.
    Poll GET /api/batch/{id} for progress.
    """
    _validate_pipeline(domain, pipeline_type)

    if len(files) > MAX_FILES_PER_BATCH:
        raise HTTPException(
            status_code=400,
            detail=f"Maximum {MAX_FILES_PER_BATCH} files per batch. Got {len(files)}."
        )

    upload_dir = Path(settings.UPLOAD_DIR) / user.id / "batch"
    upload_dir.mkdir(parents=True, exist_ok=True)

    saved_files = _extract_files_from_upload(files, upload_dir)

    if not saved_files:
        raise HTTPException(
            status_code=400,
            detail="No processable files found. Allowed: PDF, JPG, PNG, WEBP, TIFF, BMP (or ZIP of these)."
        )

    if len(saved_files) > MAX_FILES_PER_BATCH:
        # Clean up extras
        for path, _ in saved_files[MAX_FILES_PER_BATCH:]:
            Path(path).unlink(missing_ok=True)
        saved_files = saved_files[:MAX_FILES_PER_BATCH]

    # Resolve domain enum
    try:
        domain_enum = AgentDomain(domain)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid domain: {domain}")

    # Create BatchJob
    batch = BatchJob(
        user_id=user.id,
        name=name or f"Batch — {pipeline_type.replace('_', ' ').title()}",
        domain=domain_enum,
        pipeline_type=pipeline_type,
        status=BatchStatus.pending,
        total_files=len(saved_files),
        user_instructions=user_instructions or None,
    )
    db.add(batch)
    await db.flush()
    await db.refresh(batch)

    # Create BatchItem per file
    for file_path, original_filename in saved_files:
        item = BatchItem(
            batch_job_id=batch.id,
            original_filename=original_filename,
            file_path=file_path,
            status=BatchStatus.pending,
        )
        db.add(item)

    await db.commit()
    await db.refresh(batch)

    # Fire background processing
    background_tasks.add_task(run_batch_job, batch.id, user.id)
    logger.info(f"[batch:{batch.id[:8]}] created files={len(saved_files)} pipeline={domain}/{pipeline_type}")

    return batch


@router.get("", response_model=BatchJobListResponse)
async def list_batch_jobs(
    page: int = 1,
    per_page: int = 20,
    status: str = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = select(BatchJob).where(BatchJob.user_id == user.id)
    cq = select(func.count(BatchJob.id)).where(BatchJob.user_id == user.id)

    if status:
        q = q.where(BatchJob.status == status)
        cq = cq.where(BatchJob.status == status)

    total = (await db.execute(cq)).scalar()
    batches = (
        await db.execute(
            q.order_by(BatchJob.created_at.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
        )
    ).scalars().all()

    return BatchJobListResponse(batches=list(batches), total=total, page=page, per_page=per_page)


@router.get("/{batch_id}", response_model=BatchJobOut)
async def get_batch_job(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(BatchJob).where(BatchJob.id == batch_id, BatchJob.user_id == user.id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch job not found")
    return batch


@router.get("/{batch_id}/results")
async def download_batch_results(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Download all successful agent results as a multi-sheet Excel workbook.
    One sheet per processed file + a Summary sheet with aggregate stats.
    """
    result = await db.execute(
        select(BatchJob).where(BatchJob.id == batch_id, BatchJob.user_id == user.id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch job not found")

    if batch.status not in (BatchStatus.completed, BatchStatus.partial):
        raise HTTPException(status_code=400, detail="Batch not yet complete")

    # Get agent runs for this batch (matched by pipeline_type + user + time window)
    from app.models.models import AgentRun
    from sqlalchemy import and_

    agent_runs_result = await db.execute(
        select(AgentRun).where(
            AgentRun.user_id == user.id,
            AgentRun.pipeline_type == batch.pipeline_type,
            AgentRun.created_at >= batch.created_at,
        ).order_by(AgentRun.created_at.asc())
    )
    runs = agent_runs_result.scalars().all()

    if not runs:
        raise HTTPException(status_code=404, detail="No results available")

    # Build Excel workbook
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="left", vertical="center")
    alt_fill = PatternFill("solid", fgColor="F3F2EF")

    def _style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["File", "Status", "Confidence", "Summary"])
    for col in range(1, 5):
        _style_header(ws_summary.cell(1, col))
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 14
    ws_summary.column_dimensions["C"].width = 12
    ws_summary.column_dimensions["D"].width = 60

    for i, run in enumerate(runs):
        ws_summary.append([
            run.original_filename or "—",
            run.status.value,
            f"{run.confidence_score}%" if run.confidence_score else "—",
            run.summary or "—",
        ])
        if i % 2 == 1:
            for col in range(1, 5):
                ws_summary.cell(i + 2, col).fill = alt_fill

    # Per-file sheets (first 20 to keep workbook manageable)
    from app.services.export_service import _flatten

    for run in runs[:20]:
        if not run.structured_result:
            continue
        sheet_name = (run.original_filename or run.id)[:28].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Field", "Value"])
        _style_header(ws.cell(1, 1))
        _style_header(ws.cell(1, 2))
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 52

        flat = _flatten(run.structured_result)
        for i, (k, v) in enumerate(flat.items()):
            ws.append([k.replace("_", " ").title(), str(v) if v is not None else "—"])
            if i % 2 == 1:
                ws.cell(i + 2, 1).fill = alt_fill
                ws.cell(i + 2, 2).fill = alt_fill

    import io
    buf = io.BytesIO()
    wb.save(buf)

    filename = f"batch_{batch_id[:8]}_{batch.pipeline_type}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.delete("/{batch_id}", status_code=204)
async def delete_batch_job(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(BatchJob).where(BatchJob.id == batch_id, BatchJob.user_id == user.id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch job not found")

    if batch.status == BatchStatus.processing:
        raise HTTPException(
            status_code=409,
            detail="Cannot delete a batch that is currently processing. Wait for completion."
        )

    await db.delete(batch)
    await db.commit()